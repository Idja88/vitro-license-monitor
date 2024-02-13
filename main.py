import sys
from datetime import datetime
import requests
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook, load_workbook

def main(url, lic_path):
    with open(lic_path) as lic_file:
        license = lic_file.read()

    get_response = requests.get(url)

    if get_response.status_code == 200:
        get_soup = bs(get_response.text, 'html.parser')
        viewstate = get_soup.find('input', {'name': '__VIEWSTATE'})['value']
        viewstategenerator = get_soup.find('input', {'name': '__VIEWSTATEGENERATOR'})['value']
        eventvalidation = get_soup.find('input', {'name': '__EVENTVALIDATION'})['value']

    payload = {
        'txtLicenseCode': license,
        'btnGetActivationData' : 'Get Activation Data',
        'chkUseHashedMachineCodes': 'on',
        '__EVENTTARGET': '',
        '__EVENTARGUMENT': '',
        '__LASTFOCUS': '',
        '__VIEWSTATE': viewstate,
        '__VIEWSTATEGENERATOR': viewstategenerator,
        '__EVENTVALIDATION': eventvalidation
    }

    post_response = requests.post(url, data=payload)

    if post_response.status_code == 200:
        post_soup = bs(post_response.text, 'html.parser')
        stat_element = post_soup.find(id="lblStat")
        current_connections = stat_element.text.strip()
        parts = current_connections.split(', ')

        data_dict = {}
        for part in parts:
            key, value = part.split(' = ')
            data_dict[key] = int(value)

        max_activations = data_dict['Maximum activations']
        used_activations = data_dict['Used activations']
        remaining_activations = data_dict['Remaining activations']

        try:
            wb = load_workbook('data.xlsx')
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet["A1"] = "Date"
            sheet["B1"] = "Maximum"
            sheet["C1"] = "Used"
            sheet["D1"] = "Remaining"

        sheet = wb.active
        row_number = sheet.max_row + 1

        sheet[f"A{row_number}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet[f"B{row_number}"] = max_activations
        sheet[f"C{row_number}"] = used_activations
        sheet[f"D{row_number}"] = remaining_activations

        wb.save('data.xlsx')

if __name__ == "__main__":
    url = sys.argv[1]
    lic_path = sys.argv[2]
    main(url, lic_path)